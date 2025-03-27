import streamlit as st
st.set_page_config(page_title="Vyhodnocení laboratorního deníku")
st.write("Streamlit import OK")
import pandas as pd
st.write("Pandas import OK")
import pdfplumber
st.write("pdfplumber import OK")
import io
st.write("io import OK")
from openpyxl import load_workbook
st.write("openpyxl import OK")
from difflib import SequenceMatcher

st.title("Vyhodnocení laboratorního deníku")

pdf_file = st.file_uploader("Nahraj laboratorní deník (PDF)", type="pdf")
xlsx_file = st.file_uploader("Nahraj soubor Klíč.xlsx", type="xlsx")

def extract_text_from_pdf(file):
    with pdfplumber.open(file) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages)

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

def contains_similar(text, keyword, threshold=0.4):
    text = text.lower()
    keyword = keyword.lower()
    if keyword in text:
        return True
    return similar(text, keyword) >= threshold

def count_matches_advanced(text, konstrukce, zkouska_raw, stanice_raw):
    st.markdown(f"---\n🔍 **Konstrukce:** `{konstrukce}`")
    st.markdown(f"🔍 **Zkoušky:** `{zkouska_raw}`")
    st.markdown(f"🔍 **Staničení:** `{stanice_raw}`")
    druhy_zk = [z.strip().lower() for z in str(zkouska_raw).split(",") if z.strip()]
    staniceni = [s.strip().lower() for s in str(stanice_raw).split(",") if s.strip()]
    match_count = 0
    for line in text.splitlines():
        line_lower = line.lower()
        konstrukce_ok = contains_similar(line, konstrukce)
        zkouska_ok = any(z in line_lower for z in druhy_zk)
        stanice_ok = any(s in line_lower for s in staniceni)
        debug_status = f"⛔ | konstrukce_ok={konstrukce_ok}, zkouska_ok={zkouska_ok}, stanice_ok={stanice_ok}"
        if konstrukce_ok and zkouska_ok and stanice_ok:
            match_count += 1
            st.markdown(f"✅ **Shoda nalezena:** `{line.strip()}`")
        else:
            st.markdown(f"{debug_status} → `{line.strip()}`")
    st.markdown(f"**Celkem nalezeno:** `{match_count}` záznamů")
    return match_count

def process_op_sheet(key_df, target_df, lab_text):
    if "D" not in target_df.columns:
        target_df["D"] = 0
    if "E" not in target_df.columns:
        target_df["E"] = ""
    for i in range(1, len(target_df)):
        row = target_df.iloc[i]
        zasyp = str(row.iloc[0])
        if pd.isna(zasyp):
            continue
        matches = key_df[key_df.iloc[:, 0] == zasyp]
        total_count = 0
        for _, mrow in matches.iterrows():
            konstrukce = mrow.get("konstrukční prvek", "")
            zkouska = mrow.get("druh zkoušky", "")
            stanice = mrow.get("staničení", "")
            if konstrukce and zkouska and stanice:
                total_count += count_matches_advanced(lab_text, konstrukce, zkouska, stanice)
        target_df.at[i, "D"] = total_count
        pozadovano = row.get("C")
        if pd.notna(pozadovano):
            target_df.at[i, "E"] = "Vyhovující" if total_count >= pozadovano else f"Chybí {abs(int(pozadovano - total_count))} zk."
    return target_df

def process_cely_objekt_sheet(key_df, target_df, lab_text):
    for i, row in target_df.iterrows():
        material = row.get("materiál")
        zkouska = row.get("druh zkoušky")
        if pd.isna(material) or pd.isna(zkouska):
            continue
        count = count_matches_advanced(lab_text, material, zkouska, "")
        target_df.at[i, "C"] = count
        pozadovano = row.get("B")
        if pd.notna(pozadovano):
            target_df.at[i, "D"] = "Vyhovující" if count >= pozadovano else f"Chybí {abs(int(pozadovano - count))} zk."
    return target_df

if pdf_file and xlsx_file:
    lab_text = extract_text_from_pdf(pdf_file)

    st.subheader("📄 Náhled textu z PDF")
    st.text("\n".join(lab_text.splitlines()[:15]))

    try:
        xlsx_bytes = xlsx_file.read()
        workbook = load_workbook(io.BytesIO(xlsx_bytes))

        def load_sheet_df(name):
            return pd.read_excel(io.BytesIO(xlsx_bytes), sheet_name=name)

        sheet_names = workbook.sheetnames

        def sheet_exists(name):
            return name in sheet_names

        op1_key = load_sheet_df("seznam zkoušek PM+LM OP1") if sheet_exists("seznam zkoušek PM+LM OP1") else pd.DataFrame()
        op2_key = load_sheet_df("seznam zkoušek PM+LM OP2") if sheet_exists("seznam zkoušek PM+LM OP2") else pd.DataFrame()
        cely_key = load_sheet_df("seznam zkoušek Celý objekt") if sheet_exists("seznam zkoušek Celý objekt") else pd.DataFrame()

        sheet_targets = [
            ("PM - OP1", op1_key),
            ("LM - OP1", op1_key),
            ("PM - OP2", op2_key),
            ("LM - OP2", op2_key),
            ("Celý objekt", cely_key),
        ]

        for sheet_name, key_df in sheet_targets:
            if sheet_exists(sheet_name) and not key_df.empty:
                df = load_sheet_df(sheet_name)
                processed = process_cely_objekt_sheet(key_df, df, lab_text) if "Celý objekt" in sheet_name else process_op_sheet(key_df, df, lab_text)
                ws = workbook[sheet_name]
                for i, row in processed.iterrows():
                    if "D" in processed.columns:
                        ws.cell(row=i+2, column=4, value=row.get("D"))
                    if "E" in processed.columns:
                        ws.cell(row=i+2, column=5, value=row.get("E"))

        output = io.BytesIO()
        workbook.save(output)

        st.success("Vyhodnocení dokončeno. Stáhni výsledný soubor níže.")
        st.download_button(
            label="📥 Stáhnout výsledný Excel",
            data=output.getvalue(),
            file_name="vyhodnoceni_vystup.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Chyba při zpracování souboru: {e}")
